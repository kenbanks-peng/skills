#!/usr/bin/env python3
"""Regenerate the jobs workbook from jobs.db.

Usage: python export_xlsx.py <jobs.db> <output.xlsx> [--search-spec <path>]
                                                     [--include-archived]

SQLite is the source of truth. This script is idempotent: it rebuilds the
entire workbook from the current DB state on every run. Any manual edits
made directly in the .xlsx will be overwritten — edit the DB instead.

Writes to a sibling temp file and atomically renames into place, so a
crashed run won't leave a half-written workbook. If the .xlsx is open in
Excel on Windows, the rename will fail with PermissionError; close the
file and re-run.

Sheets produced:
  - Jobs     : every non-archived row (or every row with --include-archived)
  - Active   : pipeline view — not archived, not in a terminal state
  - Summary  : counts by track x state

Dependency: openpyxl. Install with:
    pip install openpyxl --break-system-packages
"""

import argparse
import os
import sqlite3
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "openpyxl is required. Install with:\n"
        "    pip install openpyxl --break-system-packages\n"
    )
    sys.exit(2)

# Re-use the search-spec parser for track labels.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from parse_search_spec import parse_search_spec  # noqa: E402
except ImportError:
    parse_search_spec = None  # type: ignore


# Columns to export, in display order. (db_column, header)
COLUMNS: list[tuple[str, str]] = [
    ("id", "ID"),
    ("date_found", "Found"),
    ("date_posted", "Posted"),
    ("track", "Track"),
    ("match_score", "Score"),
    ("priority", "Priority"),
    ("excitement", "Excitement"),
    ("state", "State"),
    ("job_title", "Title"),
    ("company", "Company"),
    ("seniority_level", "Seniority"),
    ("role_type", "Role type"),
    ("location", "Location"),
    ("work_mode", "Work mode"),
    ("employment_type", "Employment"),
    ("salary_raw", "Salary"),
    ("salary_min", "Salary min"),
    ("salary_max", "Salary max"),
    ("salary_currency", "Currency"),
    ("keyword_alignment", "Matched keywords"),
    ("match_reasons", "Why this score"),
    ("source", "Source"),
    ("job_url", "Job URL"),
    ("application_url", "Apply URL"),
    ("date_applied", "Applied"),
    ("application_method", "Method"),
    ("resume_version", "Resume"),
    ("next_action", "Next action"),
    ("next_action_date", "Next action date"),
    ("interview_count", "Interviews"),
    ("rejection_reason", "Rejection reason"),
    ("notes", "Notes"),
]

TERMINAL_STATES = ("rejected", "withdrawn", "cancelled")

HEADER_FILL = PatternFill(
    start_color="FF1F2937", end_color="FF1F2937", fill_type="solid"
)
HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def load_track_labels(spec_path):
    labels = {"unknown": "Unclassified", None: "Unclassified"}
    if spec_path and parse_search_spec is not None:
        try:
            spec = parse_search_spec(Path(spec_path).read_text(encoding="utf-8"))
            for t in spec.get("track", []):
                labels[t["id"]] = t.get("label", t["id"])
        except Exception:
            # Never let a spec problem break the export.
            pass
    return labels


def fetch_rows(conn, include_archived):
    cols = ", ".join(c for c, _ in COLUMNS) + ", archived"
    where = "" if include_archived else "WHERE archived = 0"
    sql = (
        f"SELECT {cols} FROM jobs {where} "
        "ORDER BY COALESCE(date_found,'') DESC, "
        "COALESCE(match_score, 0) DESC, id DESC"
    )
    return list(conn.execute(sql))


def write_sheet(ws, rows, track_labels):
    headers = [h for _, h in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in rows:
        out = []
        for col, _ in COLUMNS:
            val = row[col]
            if col == "track":
                val = track_labels.get(val, val or "Unclassified")
            out.append(val)
        ws.append(out)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"

    # Reasonable column widths without scanning every cell.
    widths = {
        "ID": 6,
        "Found": 11,
        "Posted": 11,
        "Track": 18,
        "Score": 7,
        "Priority": 9,
        "Excitement": 11,
        "State": 13,
        "Title": 38,
        "Company": 24,
        "Seniority": 12,
        "Role type": 10,
        "Location": 24,
        "Work mode": 10,
        "Employment": 12,
        "Salary": 22,
        "Salary min": 11,
        "Salary max": 11,
        "Currency": 9,
        "Matched keywords": 32,
        "Why this score": 40,
        "Source": 14,
        "Job URL": 48,
        "Apply URL": 48,
        "Applied": 11,
        "Method": 12,
        "Resume": 16,
        "Next action": 28,
        "Next action date": 15,
        "Interviews": 10,
        "Rejection reason": 28,
        "Notes": 40,
    }
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 14)


def write_summary(ws, conn, track_labels):
    ws.append(["Track", "State", "Count"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    rows = conn.execute(
        "SELECT track, state, COUNT(*) "
        "FROM jobs WHERE archived = 0 "
        "GROUP BY track, state "
        "ORDER BY track, state"
    ).fetchall()
    for track, state, n in rows:
        ws.append([track_labels.get(track, track or "Unclassified"), state, n])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("db", type=Path)
    ap.add_argument("output", type=Path)
    ap.add_argument("--search-spec", type=Path, default=None)
    ap.add_argument("--include-archived", action="store_true")
    args = ap.parse_args()

    if not args.db.exists():
        sys.stderr.write(f"DB not found: {args.db}\n")
        return 1

    args.output.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(str(args.db))
    conn.row_factory = sqlite3.Row
    try:
        track_labels = load_track_labels(args.search_spec)
        all_rows = fetch_rows(conn, args.include_archived)
        active_rows = [r for r in all_rows if (r["state"] or "") not in TERMINAL_STATES]

        wb = Workbook()
        jobs_sheet = wb.active
        jobs_sheet.title = "Jobs"
        write_sheet(jobs_sheet, all_rows, track_labels)

        active_sheet = wb.create_sheet("Active")
        write_sheet(active_sheet, active_rows, track_labels)

        summary_sheet = wb.create_sheet("Summary")
        write_summary(summary_sheet, conn, track_labels)

        # Atomic write: build into a temp sibling, then replace.
        tmp = args.output.with_suffix(args.output.suffix + ".tmp")
        wb.save(str(tmp))
        os.replace(tmp, args.output)
    finally:
        conn.close()

    print(f"Wrote {args.output} - {len(all_rows)} rows ({len(active_rows)} active)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
